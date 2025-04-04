import os
import pandas as pd
import json
from datetime import datetime
from channels import channels
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
import calendar
from unit import Unit
from rules import check_missing_rows
import warnings
warnings.filterwarnings(
    "ignore",
    message="This pattern is interpreted as a regular expression, and has match groups.*",
    category=UserWarning
)

block_1 = [2804, 2806, 2808, 2810, 2812, 2814, 2816, 2818]
block_3 = [77, 78, 79, 80, 81, 82, 83, 84, 85, 86]

class QualityChecker:
    def __init__(self, config_path='config/'):
        self.units = self._load_units(config_path)
        self.red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        self.yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
        
    def _load_units(self, config_path: str) -> list[Unit]:
        '''
        Load unit config jsons from folder path

        param: config_path: str: path to the config file
        return: list[Unit]: list of units
        '''
        units = []
        for file in os.listdir(config_path):
            if file.endswith('.json'):
                with open(os.path.join(config_path, file), 'r') as f:
                    unit = json.load(f)
                    units.append(Unit(unit['unit_no'], unit['block'], unit['ip_address'], unit['port'], unit['serial'], unit['channels']))
        return sorted(units)

    def _load_quality_report(self, unit, path=""):
        '''
        Find the quality report for a specific unit

        param: unit: str: unit number
        return: str: path to the quality report
        '''
        unit_no = unit.unit_no
        unit = [unit for unit in self.units if unit.unit_no == unit_no][0]
        monitored_channels = [channel for channel, key in unit.channels.items() if key == True]
        bad_df_daily = pd.DataFrame(columns=monitored_channels)
        missing_df_daily = pd.DataFrame(columns=monitored_channels)
        bad_df_monthly = pd.DataFrame(columns=monitored_channels)
        missing_df_monthly = pd.DataFrame(columns=monitored_channels)
        if path != "" and os.path.exists(path):
            try:
                with pd.ExcelFile(path) as xls:
                    if 'Daily Bad Values' in xls.sheet_names:
                        bad_df_daily = pd.read_excel(xls, 'Daily Bad Values', index_col=0)
                        bad_df_daily.index = pd.to_datetime(bad_df_daily.index, errors='coerce')
                        bad_df_daily.index = bad_df_daily.index.strftime('%Y-%m-%d')
                    if 'Daily Missing Values' in xls.sheet_names:
                        missing_df_daily = pd.read_excel(xls, 'Daily Missing Values', index_col=0)
                        missing_df_daily.index = pd.to_datetime(missing_df_daily.index, errors='coerce')
                        missing_df_daily.index = missing_df_daily.index.strftime('%Y-%m-%d')
                    if 'Monthly Bad Values' in xls.sheet_names:
                        bad_df_monthly = pd.read_excel(xls, 'Monthly Bad Values', index_col=0)
                        bad_df_monthly.index = pd.to_datetime(bad_df_monthly.index, format='%Y-%m', errors='coerce')
                        bad_df_monthly.index = bad_df_monthly.index.strftime('%Y-%m')
                    if 'Monthly Missing Values' in xls.sheet_names:
                        missing_df_monthly = pd.read_excel(xls, 'Monthly Missing Values', index_col=0)
                        missing_df_monthly.index = pd.to_datetime(missing_df_monthly.index, format='%Y-%m', errors='coerce')
                        missing_df_monthly.index = missing_df_monthly.index.strftime('%Y-%m')
            except Exception as e:
                print(f"Error reading existing report: {str(e)}")
        return (bad_df_daily, missing_df_daily, bad_df_monthly, missing_df_monthly)

    def check_data_quality(self, unit_no):
        '''
        Check the quality of the data
        '''
        unit = [unit for unit in self.units if unit.unit_no == unit_no][0]
        bad_df_daily, missing_df_daily, bad_df_monthly, missing_df_monthly = self._load_quality_report(unit, f'quality_reports/UNIT {unit_no} REPORT.xlsx')
        # Get list of dates from unit.data
        if not unit.load_data(f'Minute_Data/'):
            print(f'Unit {unit.unit_no} has no data')
            return ((bad_df_daily, missing_df_daily), (bad_df_monthly, missing_df_monthly))
        
        unit.data['Date'] = pd.to_datetime(unit.data['Date'], errors='coerce')
        unit.data = unit.data.dropna(subset=['Date'])
        unit.data, _, _, _ = check_missing_rows(unit.data, unit.unit_no)
        if 'Date' in unit.data.columns:
            unit.data = unit.data.loc[:, ~unit.data.columns.duplicated()]
        unit.data.set_index('Date', inplace=True)
        data = unit.data
        data = data.sort_index()
        # Group data by date (each group corresponds to one day)
        daily_groups = data.groupby(data.index.date)
        # Extract unique year-month combinations in the format: YYYY-MM
        unique_months = np.unique([date.strftime('%Y-%m') for date, _ in daily_groups])
        monitored_channels = [channel for channel, key in unit.channels.items() if key == True]
        # Process daily quality checks using vectorized operations
        for date, daily_data in daily_groups:
            date_str = date.strftime('%Y-%m-%d') 
            for channel in monitored_channels:
                min_val = channels[channel].min_value
                max_val = channels[channel].max_value
                
                # Find the column matching the channel regex
                matching_cols = data.columns[data.columns.str.contains(channels[channel].regex, regex=True)]
                channel_name = matching_cols[0] if len(matching_cols) > 0 else None
                if channel_name is None:
                    continue

                # Create vectorized masks for bad and missing values
                col_numeric = pd.to_numeric(daily_data[channel_name], errors='coerce')
                bad_mask = (col_numeric < min_val) | (col_numeric > max_val)
                missing_mask = col_numeric.isna() | np.isinf(col_numeric)
                bad_values = bad_mask.sum()
                missing_values = missing_mask.sum()

                # Compute percentage assuming an expected 1440 data points per day
                bad_df_daily.loc[date_str, channel] = float(round(float(bad_values) / 1440 * 100, 3))
                missing_df_daily.loc[date_str, channel] = float(round(float(missing_values) / 1440 * 100, 3))
                    
        daily = (bad_df_daily, missing_df_daily)
        for month in unique_months:
            year = int(month.split('-')[0])
            month_num = int(month.split('-')[1])

            num_days = calendar.monthrange(year, month_num)[1]
            # days_list = missing_df_daily.index.str.startswith(month)
            # actual_num_days = len(missing_df_daily[days_list])
            # missing_num_days = num_days - actual_num_days
            expected_num_points = 1440*num_days
            # Filter data for the current month
            month_str = month  # month is already in 'YYYY-MM' format from unique_months
            monthly_data = data[data.index.strftime('%Y-%m') == month_str]

            # Process data for each enabled channel
            for channel in monitored_channels:
                min_val = channels[channel].min_value
                max_val = channels[channel].max_value


                # Find the column matching the channel regex
                matching_cols = data.columns[data.columns.str.contains(channels[channel].regex, regex=True)]
                channel_name = matching_cols[0] if len(matching_cols) > 0 else None
                if channel_name is None:
                    continue

                # Create vectorized masks for bad and missing values
                col_numeric = pd.to_numeric(monthly_data[channel_name], errors='coerce')
                bad_mask = (col_numeric < min_val) | (col_numeric > max_val)
                missing_mask = col_numeric.isna() | np.isinf(col_numeric)
                bad_values = bad_mask.sum()
                missing_values = missing_mask.sum()

                # Compute percentage using the expected number of data points
                bad_df_monthly.loc[month_str, channel] = float(round((bad_values)/ (expected_num_points) * 100, 3))
                missing_num_points = (missing_values + expected_num_points - len(col_numeric))
                missing_df_monthly.loc[month_str, channel] = float(round(missing_num_points/ (expected_num_points) * 100, 3))
                monthly = (bad_df_monthly, missing_df_monthly)
                if month_str == '2025-03':
                        pass
        # Add new derived columns to bad_df and missing_df
        for df in [bad_df_monthly, missing_df_monthly]:
            # Sum Main Electricity 1 & 2 with percentage calculation
            elec_cols = df.columns[df.columns.str.contains('Main\\s*Electricity\\s*[12]\\s*(Watts)$', regex=True)]
            if len(elec_cols) >= 2:
                df['Total Electricity'] = df[elec_cols].sum(axis=1)/len(elec_cols)
            
            # Copy Natural Gas column
            if 'Natural Gas' in df.columns:
                df['Gas'] = df['Natural Gas']

        return (daily, monthly)

    def _format_quality_result(self, path):
        wb = load_workbook(path)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Skip sheets with no data
            if ws.max_row <= 1 or ws.max_column <= 1:
                continue
                
            # Get the maximum row and column
            max_row = ws.max_row
            max_col = ws.max_column
            
            # Create range string for formatting (B2 to end)
            end_col_letter = get_column_letter(max_col)
            data_range = f"B2:{end_col_letter}{max_row}"
            
            try:
                # Create a color scale rule
                color_scale = ColorScaleRule(
                    start_type="num", start_value=0, start_color="FFFFFFFF",  # White for 0
                    end_type="num", end_value=100, end_color="FFFF0000"      # Red for 100
                )
                
                # Apply the color scale to the data range
                ws.conditional_formatting.add(data_range, color_scale)
            except Exception as e:
                print(f"Warning: Could not apply formatting to sheet '{sheet_name}': {str(e)}")
                
        wb.save(path)

    def update_quality_report(self, unit_no, dataframes):
        '''
        Update the quality report with the new data
        '''
        daily, monthly = dataframes
        if not os.path.exists(f'quality_reports/'):
            os.makedirs(f'quality_reports/')
        path = f'quality_reports/UNIT {unit_no} REPORT.xlsx'
        with pd.ExcelWriter(path) as writer:
            bad_df_daily, missing_df_daily = daily
            bad_df_daily.to_excel(writer, sheet_name='Daily Bad Values')
            missing_df_daily.to_excel(writer, sheet_name='Daily Missing Values')
            bad_df_monthly, missing_df_monthly = monthly
            bad_df_monthly.to_excel(writer, sheet_name='Monthly Bad Values')
            missing_df_monthly.to_excel(writer, sheet_name='Monthly Missing Values')
        try:
            self._format_quality_result(path)
        except Exception as e:
            print(f"Error formatting quality report: {str(e)}")
        print(f'Quality report for unit {unit_no} updated')

        return
    
def main():
    checker = QualityChecker()
    for unit in block_1+block_3:
        dataframes = checker.check_data_quality(unit)
        checker.update_quality_report(unit, dataframes)


if __name__ == "__main__":
    # checker = QualityChecker()
    # dataframes = checker.check_data_quality(2812)
    # checker.update_quality_report(2812, dataframes)
    main()
